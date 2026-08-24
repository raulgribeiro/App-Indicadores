# -*- coding: utf-8 -*-
"""
update_dashboard.py
====================
Lê as planilhas Agrícola.xlsx e Industria.xlsx da pasta do OneDrive,
regenera os blocos de dados do index.html (dashboard) e publica a
atualização no GitHub (add + commit + push).

Como funciona (visão geral):
  1. Lê as duas planilhas com openpyxl.
  2. Interpreta a estrutura de cada bloco de indicador dinamicamente
     (não depende de número de linha fixo — se os VALORES mudarem na
     planilha, o script pega os valores novos automaticamente).
  3. Usa duas tabelas de metadados (AGRICOLA_META / INDUSTRIA_META)
     para saber o nome bonito, categoria e unidade de cada indicador.
     Se aparecer um indicador novo que não está nessas tabelas, o
     script AVISA no terminal em vez de adivinhar errado — é só
     adicionar uma linha na tabela correspondente.
  4. Substitui o conteúdo entre os marcadores
     // ==DADOS_AGRICOLA_START== ... // ==DADOS_AGRICOLA_END==
     // ==DADOS_INDUSTRIA_START== ... // ==DADOS_INDUSTRIA_END==
     dentro do index.html do repositório.
  5. Roda git add / commit / push no repositório.

Só rode dando duplo-clique em "update_dashboard.bat" (que chama este
script), ou manualmente com: python update_dashboard.py
"""

import datetime
import json
import re
import subprocess
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERRO: falta o pacote 'openpyxl'.")
    print("Instale com:  pip install openpyxl")
    sys.exit(1)

# ============================================================
# CONFIGURAÇÃO — ajuste aqui se algo mudar
# ============================================================

# Pasta do OneDrive onde ficam as planilhas de origem
PASTA_PLANILHAS = Path(
    r"C:\Users\raulribeiro\OneDrive - CLEALCO AÇÚCAR E ÁLCOOL S.A"
    r"\Projeto Confiar Excelência\Excelência em Processo\Governanças\Pasta Indicadores"
)
ARQ_AGRICOLA = PASTA_PLANILHAS / "Agrícola.xlsx"
ARQ_INDUSTRIA = PASTA_PLANILHAS / "Industria.xlsx"

# Pasta do repositório git local (onde está o index.html publicado no GitHub Pages).
# Por padrão, assume que este script está DENTRO da pasta do repositório.
REPO_DIR = Path(__file__).resolve().parent
INDEX_HTML = REPO_DIR / "index.html"

# Mensagem de commit
def mensagem_commit():
    agora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    return f"Atualização automática de indicadores — {agora}"


# ============================================================
# METADADOS — nome bonito / categoria / unidade / direção de cada indicador.
# Os VALORES vêm da planilha automaticamente; isso aqui é só a "ficha"
# de cada indicador, que raramente muda. Se um indicador novo aparecer
# na planilha e não estiver aqui, o script avisa em vez de adivinhar.
# ============================================================

# chave = (TÍTULO EM MAIÚSCULO como aparece na planilha, marcador ou None)
AGRICOLA_META = {
    ("EFICIÊNCIA OPERACIONAL COLHEITA", None): dict(id="colheita_op", categoria="Colheita", nome="Eficiência Operacional Colheita", unidade="%", lowerIsBetter=False),
    ("EFICIÊNCIA MOTOR COLHEITA", None): dict(id="colheita_motor", categoria="Colheita", nome="Eficiência Motor Colheita", unidade="%", lowerIsBetter=False),
    ("EFICIÊNCIA OPERACIONAL TRANSPORTE", None): dict(id="transporte", categoria="Transporte", nome="Eficiência Operacional Transporte", unidade="%", lowerIsBetter=False),
    ("EFICIÊNCIA OPERACIONAL VINHAÇA LOCALIZADA", None): dict(id="vinhaca_op", categoria="Vinhaça Localizada", nome="Eficiência Operacional Vinhaça Localizada", unidade="%", lowerIsBetter=False),
    ("EFICIÊNCIA MOTOR VINHAÇA LOCALIZADA", None): dict(id="vinhaca_motor", categoria="Vinhaça Localizada", nome="Eficiência Motor Vinhaça Localizada", unidade="%", lowerIsBetter=False),
    ("EFICIÊNCIA OPERACIONAL RODOVINHAÇA", None): dict(id="rodovinhaca", categoria="Rodovinhaça", nome="Eficiência Operacional Rodovinhaça", unidade="%", lowerIsBetter=False),
    ("EFICIÊNCIA OPERACIONAL PLANTIO", "F1"): dict(id="plantio_op_f1", categoria="Plantio", nome="Eficiência Operacional Plantio — F1", unidade="%", lowerIsBetter=False),
    ("EFICIÊNCIA OPERACIONAL PLANTIO", "F2"): dict(id="plantio_op_f2", categoria="Plantio", nome="Eficiência Operacional Plantio — F2", unidade="%", lowerIsBetter=False),
    ("EFICIÊNCIA OPERACIONAL PLANTIO", "MÉDIA"): dict(id="plantio_op_media", categoria="Plantio", nome="Eficiência Operacional Plantio — Média Geral", unidade="%", lowerIsBetter=False),
    ("EFICIÊNCIA MOTOR PLANTIO", "F1"): dict(id="plantio_motor_f1", categoria="Plantio", nome="Eficiência Motor Plantio — F1", unidade="%", lowerIsBetter=False),
    ("EFICIÊNCIA MOTOR PLANTIO", "F2"): dict(id="plantio_motor_f2", categoria="Plantio", nome="Eficiência Motor Plantio — F2", unidade="%", lowerIsBetter=False),
    ("EFICIÊNCIA MOTOR PLANTIO", "MÉDIA"): dict(id="plantio_motor_media", categoria="Plantio", nome="Eficiência Motor Plantio — Média Geral", unidade="%", lowerIsBetter=False),
    ("TEMPO DE MANOBRA", None): dict(id="manobra", categoria="Frota e Consumo", nome="Tempo de Manobra", unidade="h", lowerIsBetter=True),
    ("REDUÇÃO DE CONSUMO CANAVIEIRO L/TON", None): dict(id="consumo_canavieiro", categoria="Frota e Consumo", nome="Redução de Consumo Canavieiro", unidade="L/TON", lowerIsBetter=True),
    ("OTIMIZAÇÃO DE PNEUS", None): dict(id="pneus", categoria="Frota e Consumo", nome="Otimização de Pneus", unidade="%", lowerIsBetter=True),
    ("REDUÇÃO DE CONSUMO COLHEDORA L/TON", None): dict(id="consumo_colhedora", categoria="Frota e Consumo", nome="Redução de Consumo Colhedora", unidade="L/TON", lowerIsBetter=True),
    ("TEMPO DE PATIO CANAVIEIRO", None): dict(id="patio", categoria="Pátio", nome="Tempo de Pátio Canavieiro", unidade="min", lowerIsBetter=True),
    ("TEMPO DE PÁTIO CANAVIEIRO", None): dict(id="patio", categoria="Pátio", nome="Tempo de Pátio Canavieiro", unidade="min", lowerIsBetter=True),
}

# chave = TÍTULO (sem prefixo "CLE - " / "QRZ - ") em maiúsculo
INDUSTRIA_META = {
    "EXTRAÇÃO ART": dict(id="extracao_art", categoria="Extração e Fermentação", nome="Extração ART"),
    "EFICIÊNCIA DESTILAÇÃO": dict(id="ind_destilacao", categoria="Extração e Fermentação", nome="Eficiência Destilação"),
    "EFICIÊNCIA FERMENTAÇÃO": dict(id="ind_fermentacao", categoria="Extração e Fermentação", nome="Eficiência Fermentação"),
    "PERDA ASPERSOR": dict(id="ind_perda_aspersor", categoria="Perdas de Processo", nome="Perda Aspersor"),
    "PERDA RESIDUÁRIA": dict(id="ind_perda_residuaria", categoria="Perdas de Processo", nome="Perda Residuária"),
    "POL DA TORTA": dict(id="ind_pol_torta", categoria="Perdas de Processo", nome="Pol da Torta"),
    "AD. CONTROLE DE PH": dict(id="ind_ad_ph", categoria="Água de Diluição (AD)", nome="AD. Controle de PH"),
    "AD. PRESSÃO DE 21": dict(id="ind_ad_pressao21", categoria="Água de Diluição (AD)", nome="AD. Pressão de 21"),
    "AD. TEMPERATURA DO 21": dict(id="ind_ad_temp21", categoria="Água de Diluição (AD)", nome="AD. Temperatura do 21"),
}

PERIODO_LABEL_MAP_INDUSTRIA = {"SF 25/26": "SF 25/26", "SEM": "SEMANA", "MÊS": "MÊS", "CONS": "ACUMULADO"}


# ============================================================
# UTILITÁRIOS
# ============================================================
def slugify(txt):
    txt = unicodedata.normalize("NFKD", txt).encode("ascii", "ignore").decode()
    txt = re.sub(r"[^a-zA-Z0-9]+", "_", txt.strip().lower()).strip("_")
    return txt


def valor_numerico(v):
    """Converte célula da planilha em número (ou None). Trata '-'/'x' como sem dado
    e horários (datetime.time) como minutos."""
    if v is None:
        return None
    if isinstance(v, datetime.time):
        return v.hour * 60 + v.minute
    if isinstance(v, str):
        vs = v.strip()
        if vs in ("-", "x", "X", ""):
            return None
        try:
            return float(vs.replace(",", "."))
        except ValueError:
            return None
    if isinstance(v, (int, float)):
        return v
    return None


def js_val(v):
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "true" if v else "false"
    return json.dumps(v, ensure_ascii=False)


# ============================================================
# PARSER — AGRÍCOLA.XLSX
# ============================================================
def parse_agricola(caminho):
    if not caminho.exists():
        raise FileNotFoundError(f"Não encontrei a planilha: {caminho}")

    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.worksheets[0]

    max_row, max_col = 0, 0
    for r in range(1, ws.max_row + 2):
        for c in range(1, ws.max_column + 2):
            if ws.cell(row=r, column=c).value is not None:
                max_row = max(max_row, r)
                max_col = max(max_col, c)

    grupos = []
    avisos = []

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None or str(v).strip().upper() != "META":
                continue

            meta_col = c
            # varre à esquerda por células de cabeçalho contíguas (nomes de série)
            series_cols = []
            cc = c - 1
            while cc >= 1:
                hv = ws.cell(row=r, column=cc).value
                if hv is None:
                    break
                series_cols.insert(0, (cc, str(hv).strip()))
                cc -= 1
            if not series_cols:
                continue

            label_col = series_cols[0][0] - 1
            if label_col < 1:
                continue

            # título: linha acima, procurando à esquerda a partir de label_col
            titulo = None
            for tc in range(label_col, 0, -1):
                tv = ws.cell(row=r - 1, column=tc).value
                if tv:
                    titulo = str(tv).strip()
                    break
            if titulo is None:
                continue

            # marcador (F1/F2/MÉDIA) — fica na primeira linha de dados, coluna 1
            marcador = None
            primeira_label = ws.cell(row=r + 1, column=label_col).value
            if primeira_label is not None:
                marca_cell = ws.cell(row=r + 1, column=1).value
                if marca_cell:
                    marcador = str(marca_cell).strip().upper()

            # linhas de dados (até encontrar rótulo vazio, máx. 8 por segurança)
            periodos = []
            dr = r + 1
            passos = 0
            while dr <= max_row and passos < 8:
                label = ws.cell(row=dr, column=label_col).value
                if label is None:
                    break
                valores = {}
                for (col, nome_serie) in series_cols:
                    valores[nome_serie] = valor_numerico(ws.cell(row=dr, column=col).value)
                meta_val = valor_numerico(ws.cell(row=dr, column=meta_col).value)
                periodos.append({"periodo": str(label).strip(), "valores": valores, "meta": meta_val})
                dr += 1
                passos += 1

            if not periodos:
                continue

            chave = (titulo.upper(), marcador)
            meta_info = AGRICOLA_META.get(chave)
            if meta_info is None:
                avisos.append(
                    f"[AGRÍCOLA] Indicador não mapeado em AGRICOLA_META: título='{titulo}' marcador={marcador!r} "
                    f"(linha {r}). Adicione uma entrada em AGRICOLA_META para ele aparecer corretamente no dashboard."
                )
                gid = slugify(titulo + ("_" + marcador if marcador else ""))
                meta_info = dict(id=gid, categoria="Outros", nome=titulo.title(), unidade="%", lowerIsBetter=False)

            grupos.append({
                "id": meta_info["id"],
                "categoria": meta_info["categoria"],
                "nome": meta_info["nome"],
                "unidade": meta_info["unidade"],
                "lowerIsBetter": meta_info["lowerIsBetter"],
                "series": [s for (_, s) in series_cols],
                "periodos": periodos,
            })

    # remove duplicados por id, mantendo o primeiro encontrado
    vistos = set()
    grupos_unicos = []
    for g in grupos:
        if g["id"] in vistos:
            continue
        vistos.add(g["id"])
        grupos_unicos.append(g)

    return grupos_unicos, avisos


def gerar_js_agricola(grupos):
    linhas = ["const grupos = ["]
    for g in grupos:
        linhas.append(
            f"  {{ id:'{g['id']}', categoria:'{g['categoria']}', nome:'{g['nome']}', "
            f"unidade:'{g['unidade']}', lowerIsBetter:{js_val(g['lowerIsBetter'])},"
        )
        linhas.append(f"    series:[{', '.join(js_val(s) for s in g['series'])}],")
        linhas.append("    periodos:[")
        for p in g["periodos"]:
            valores_txt = ", ".join(f"{k}:{js_val(v)}" for k, v in p["valores"].items())
            linhas.append(f"      {{periodo:{js_val(p['periodo'])}, valores:{{{valores_txt}}}, meta:{js_val(p['meta'])}}},")
        linhas.append("    ]},")
    linhas.append("];")
    return "\n".join(linhas)


# ============================================================
# PARSER — INDUSTRIA.XLSX (abas CLE e QRZ)
# ============================================================
def parse_industria_sheet(ws):
    max_row = 0
    for r in range(1, ws.max_row + 2):
        for c in range(1, 13):
            if ws.cell(row=r, column=c).value is not None:
                max_row = max(max_row, r)

    grupos = []
    for r in range(1, max_row + 1):
        for c in range(1, 13):
            v = ws.cell(row=r, column=c).value
            if v != "RESULTADO":
                continue

            title_row = r - 1
            titulo = None
            for tc in range(c - 1, 0, -1):
                tv = ws.cell(row=title_row, column=tc).value
                if tv:
                    titulo = tv
                    break
            if titulo is None:
                continue

            headers = {}
            cc = c
            while cc <= c + 4:
                hv = ws.cell(row=r, column=cc).value
                if hv:
                    headers[str(hv).strip().rstrip(".")] = cc
                cc += 1
                if hv and "DIRE" in str(hv):
                    break

            label_col = c - 1
            periodos = []
            for dr in range(r + 1, r + 5):
                label = ws.cell(row=dr, column=label_col).value
                if label is None:
                    label = ws.cell(row=dr, column=label_col - 1).value

                def getv(nome):
                    col = headers.get(nome)
                    return ws.cell(row=dr, column=col).value if col else None

                resultado = valor_numerico(getv("RESULTADO"))
                meta_mes = getv("META MÊS")
                if meta_mes is None:
                    meta_mes = getv("META")
                meta_mes = valor_numerico(meta_mes)
                meta_cons = valor_numerico(getv("META CONS"))
                direcao = getv("DIREÇÃO")

                periodos.append({
                    "periodo": (str(label).strip() if label else ""),
                    "valor": resultado,
                    "metaMes": meta_mes,
                    "metaCons": meta_cons,
                    "direcao": direcao,
                })

            grupos.append({"titulo": str(titulo).strip(), "periodos": periodos})
    return grupos


def parse_industria(caminho):
    if not caminho.exists():
        raise FileNotFoundError(f"Não encontrei a planilha: {caminho}")

    wb = openpyxl.load_workbook(caminho, data_only=True)
    if "CLE" not in wb.sheetnames or "QRZ" not in wb.sheetnames:
        raise ValueError(f"Esperava abas 'CLE' e 'QRZ' em {caminho.name}, encontrei: {wb.sheetnames}")

    cle = parse_industria_sheet(wb["CLE"])
    qrz = parse_industria_sheet(wb["QRZ"])

    def norm_title(t):
        return re.sub(r"^(CLE|QRZ)\s*-\s*", "", t).strip()

    cle_map = {norm_title(g["titulo"]): g for g in cle}
    qrz_map = {norm_title(g["titulo"]): g for g in qrz}

    ordem = list(cle_map.keys())
    avisos = []
    resultado = []

    for titulo in ordem:
        gcle = cle_map[titulo]
        gqrz = qrz_map.get(titulo)

        direcao = None
        for p in gcle["periodos"]:
            if p["direcao"]:
                direcao = p["direcao"]
                break
        if direcao is None and gqrz:
            for p in gqrz["periodos"]:
                if p["direcao"]:
                    direcao = p["direcao"]
                    break
        lower_is_better = (direcao == "MENOR MELHOR")

        meta_info = INDUSTRIA_META.get(titulo.upper())
        if meta_info is None:
            avisos.append(
                f"[INDÚSTRIA] Indicador não mapeado em INDUSTRIA_META: título='{titulo}'. "
                f"Adicione uma entrada em INDUSTRIA_META para ele aparecer corretamente no dashboard."
            )
            meta_info = dict(id="ind_auto_" + slugify(titulo), categoria="Outros", nome=titulo.title())

        periodos_out = []
        for i, p_cle in enumerate(gcle["periodos"]):
            p_qrz = gqrz["periodos"][i] if gqrz and i < len(gqrz["periodos"]) else {"valor": None, "metaMes": None, "metaCons": None}
            periodo_label = PERIODO_LABEL_MAP_INDUSTRIA.get(p_cle["periodo"], p_cle["periodo"])
            periodos_out.append({
                "periodo": periodo_label,
                "CLE": p_cle["valor"], "QRZ": p_qrz["valor"],
                "metaMesCLE": p_cle["metaMes"], "metaConsCLE": p_cle["metaCons"],
                "metaMesQRZ": p_qrz["metaMes"], "metaConsQRZ": p_qrz["metaCons"],
            })

        resultado.append({
            "id": meta_info["id"],
            "categoria": meta_info["categoria"],
            "nome": meta_info["nome"],
            "unidade": "%",
            "lowerIsBetter": lower_is_better,
            "periodos": periodos_out,
        })

    return resultado, avisos


def gerar_js_industria(grupos):
    linhas = ["const gruposIndustria = ["]
    for g in grupos:
        linhas.append(
            f"  {{ id:'{g['id']}', categoria:'{g['categoria']}', nome:'{g['nome']}', "
            f"unidade:'{g['unidade']}', lowerIsBetter:{js_val(g['lowerIsBetter'])},"
        )
        linhas.append("    periodos:[")
        for p in g["periodos"]:
            linhas.append(
                f"      {{periodo:{js_val(p['periodo'])}, "
                f"valores:{{CLE:{js_val(p['CLE'])}, QRZ:{js_val(p['QRZ'])}}}, "
                f"metaMes:{{CLE:{js_val(p['metaMesCLE'])}, QRZ:{js_val(p['metaMesQRZ'])}}}, "
                f"metaCons:{{CLE:{js_val(p['metaConsCLE'])}, QRZ:{js_val(p['metaConsQRZ'])}}}}},"
            )
        linhas.append("    ]},")
    linhas.append("];")
    return "\n".join(linhas)


# ============================================================
# ATUALIZAÇÃO DO index.html
# ============================================================
def substituir_bloco(html, inicio_marca, fim_marca, novo_bloco):
    padrao = re.compile(
        re.escape(inicio_marca) + r".*?" + re.escape(fim_marca), re.S
    )
    if not padrao.search(html):
        raise RuntimeError(f"Não encontrei os marcadores {inicio_marca} / {fim_marca} no index.html")
    substituto = f"{inicio_marca}\n{novo_bloco}\n{fim_marca}"
    return padrao.sub(lambda m: substituto, html, count=1)


def atualizar_index_html(agricola_js, industria_js):
    if not INDEX_HTML.exists():
        raise FileNotFoundError(
            f"Não encontrei {INDEX_HTML}. Este script precisa estar na mesma pasta do index.html do repositório."
        )
    html = INDEX_HTML.read_text(encoding="utf-8")
    html = substituir_bloco(html, "// ==DADOS_AGRICOLA_START==", "// ==DADOS_AGRICOLA_END==", agricola_js)
    html = substituir_bloco(html, "// ==DADOS_INDUSTRIA_START==", "// ==DADOS_INDUSTRIA_END==", industria_js)
    INDEX_HTML.write_text(html, encoding="utf-8")


# ============================================================
# GIT — commit e push
# ============================================================
def rodar_git(args):
    resultado = subprocess.run(
        ["git"] + args, cwd=str(REPO_DIR), capture_output=True, text=True
    )
    return resultado


def publicar_no_github():
    if not (REPO_DIR / ".git").exists():
        print(f"AVISO: {REPO_DIR} não parece ser um repositório git (pasta .git não encontrada).")
        print("Rode 'git init' / 'git clone' antes, ou ajuste REPO_DIR no topo do script.")
        return False

    diff = rodar_git(["status", "--porcelain", "index.html"])
    if diff.returncode != 0:
        print("ERRO ao checar status do git:", diff.stderr)
        return False
    if not diff.stdout.strip():
        print("Nenhuma alteração nos dados — index.html já está igual. Nada para publicar.")
        return True

    add = rodar_git(["add", "index.html"])
    if add.returncode != 0:
        print("ERRO no 'git add':", add.stderr)
        return False

    commit = rodar_git(["commit", "-m", mensagem_commit()])
    if commit.returncode != 0:
        print("ERRO no 'git commit':", commit.stderr)
        return False
    print(commit.stdout)

    push = rodar_git(["push"])
    if push.returncode != 0:
        print("ERRO no 'git push':", push.stderr)
        print("Verifique se o remoto está configurado e se você tem permissão (git remote -v).")
        return False
    print(push.stdout or push.stderr)
    print("Publicado no GitHub com sucesso.")
    return True


# ============================================================
# MAIN
# ============================================================
def main():
    print("=" * 60)
    print("Atualização do Dashboard de Indicadores")
    print("=" * 60)

    print(f"\nLendo: {ARQ_AGRICOLA}")
    grupos_agro, avisos_agro = parse_agricola(ARQ_AGRICOLA)
    print(f"  -> {len(grupos_agro)} indicadores encontrados (Agrícola).")

    print(f"\nLendo: {ARQ_INDUSTRIA}")
    grupos_ind, avisos_ind = parse_industria(ARQ_INDUSTRIA)
    print(f"  -> {len(grupos_ind)} indicadores encontrados (Indústria).")

    avisos = avisos_agro + avisos_ind
    if avisos:
        print("\n--- AVISOS (indicadores fora das tabelas de metadados) ---")
        for a in avisos:
            print("  " + a)
        print("O dashboard ainda foi gerado, usando nome/categoria padrão para esses casos.")

    print("\nGerando blocos de dados...")
    agricola_js = gerar_js_agricola(grupos_agro)
    industria_js = gerar_js_industria(grupos_ind)

    print(f"Atualizando {INDEX_HTML} ...")
    atualizar_index_html(agricola_js, industria_js)
    print("index.html atualizado.")

    print("\nPublicando no GitHub...")
    publicar_no_github()

    print("\nConcluído.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("\nERRO:", e)
        sys.exit(1)
